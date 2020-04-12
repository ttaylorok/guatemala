from openpyxl import load_workbook
import mysql.connector
wb = load_workbook(filename='2018_crime_NP_agreviado.xlsx', read_only=True)
#wb = load_workbook(filename='/Users/comp/Downloads/2018_crime_main_results_MP.xls', read_only=True)
#wb = load_workbook(filename='2018_crime_MP_dictionary.xlsx', read_only=True)

print(wb.sheetnames)
ws = wb['Sheet1']



##
##field_lengths = [0, 0, 0]
##description = ""
for row in ws.rows:
    #print(row)
    row_array = []
    for cell in row:
        row_array.append(cell.value)
    print(row_array)
