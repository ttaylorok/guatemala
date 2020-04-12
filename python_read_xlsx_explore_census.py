from openpyxl import load_workbook

wb = load_workbook(filename='db_csv/Diccionario de datos/Diccionario_Base_PERSONA.xlsx', read_only=True)
#wb = load_workbook(filename='db_csv/Diccionario de datos/Diccionario_Base_HOGAR.xlsx', read_only=True)

print(wb.sheetnames)
ws = wb['PERSONA']

for row in ws.rows:
    row_array = []
    for cell in row:
        row_array.append(cell.value)
    print(row_array)
