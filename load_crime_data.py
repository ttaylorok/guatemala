from openpyxl import load_workbook
import mysql.connector
wb = load_workbook(filename='2018_crime_NP_agreviado.xlsx', read_only=True)
#wb = load_workbook(filename='/Users/comp/Downloads/2018_crime_main_results_MP.xls', read_only=True)
#wb = load_workbook(filename='2018_crime_MP_dictionary.xlsx', read_only=True)

print(wb.sheetnames)
ws = wb['Sheet1']

mydb = mysql.connector.connect(
# REMOVED
)

mycursor = mydb.cursor()

mycursor.execute("USE guatemala")
##
##field_lengths = [0, 0, 0]
##description = ""
for row in ws.rows:
    #print(row)
    row_array = []
    for cell in row:
        row_array.append(cell.value)
##    if(row_array[0] != None):
##        description = row_array[0]
##    row_array[0] = description
##    print(row_array)
##
####    # find value ranges
####    if(len(row_array[0]) > field_lengths[0]):
####        field_lengths[0] = len(row_array[0])
####    try:
####        if(row_array[1] > field_lengths[1]):
####            field_lengths[1] = row_array[1]
####    except:
####        pass
####    try:
####        if(len(row_array[2]) > field_lengths[2]):
####            field_lengths[2] = len(row_array[2])
####    except:
####        pass
##    
    try:
        #sql = "INSERT INTO crime_stats VALUES (%d, %d, %d, %d, %d, %d, %d, %d, %d, %d, %d, %d, %d, %d, %d, %d, %d)" % (row_array[0], row_array[1], row_array[2], row_array[3], row_array[4], row_array[5], row_array[6], row_array[7], row_array[8], row_array[9], row_array[10], row_array[11], row_array[12], row_array[13], row_array[14], row_array[15], row_array[16])
        sql = """INSERT INTO crime_stats 
        (num_corre, ano_denuncia, mes_denuncia, dia_denuncia, 
        dia_sem_denuncia, ano_hecho, mes_hecho, dia_hecho, dia_sem_hecho, 
        depto_ocu, mupio_ocu, zona_ocu, sexo_agraviados, edad_agrav,
        g_edad, g_edad_2, est_conyugal, delito_com) 
        VALUES (%d, %d, %d, %d, %d, %d, %d, %d, %d, %d, %d, %d, %d, %d, %d, %d, %d, %d)""" %  (row_array[0], row_array[1], row_array[2], row_array[3], row_array[4], row_array[5], row_array[6], row_array[7], row_array[8], row_array[9], row_array[10], row_array[11], row_array[12], row_array[13], row_array[14], row_array[15], row_array[16], row_array[17])
        #print(sql)
        mycursor.execute(sql)
        mydb.commit()
    except:
        #pass
        print(row_array)
    

#mydb.commit()
mycursor.close()
mydb.close()

#print(field_lengths)
        

